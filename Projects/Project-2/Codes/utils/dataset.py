"""
dataset.py
"""

import os
import re
from math import cos, radians, sin, sqrt
from typing import List, Union

import matplotlib.pyplot as plt
import networkx as nx
import numpy as np
import pandas as pd
import scipy.cluster.hierarchy as sch
import seaborn as sns
from openpyxl import load_workbook
from rich.columns import Columns
from rich.console import Console
from scipy.stats import f_oneway
from sklearn.manifold import MDS
from sklearn.metrics.pairwise import cosine_similarity, euclidean_distances
from sklearn.preprocessing import minmax_scale


class DatasetLoaderXL:
    """
    A class to load datasets from Excel files into lists of lists.
    Using openpyxl to load Excel files.
    """

    def __init__(
        self,
        dataset_dir: Union[str, os.PathLike],
        seed=42,
        console=Console(),
        verbose: bool = False,
    ) -> None:
        self.dataset_dir = dataset_dir
        self.seed = seed
        self.console = console
        self.verbose = verbose

        self.dataset = {}

    def get_verbose(self, verbose):
        if verbose is None:
            return self.verbose
        else:
            return verbose

    def load_dataset(self, file_path) -> List[List]:
        """Load a single Excel file"""
        try:
            workbook = load_workbook(file_path)
            sheet = workbook.active
            data = [row for row in sheet.iter_rows(values_only=True)]
            return data
        except Exception as e:
            print(f"Error loading {file_path}: {e}")
            return []

    def load_all_datasets(self) -> None:
        """Load all the Excel files using openpyxl."""
        if self.dataset:
            return

        for filename in os.listdir(self.dataset_dir):
            if filename.endswith(".xlsx"):
                file_path = os.path.join(self.dataset_dir, filename)
                suburb_name = re.search(r"(.+)-Suburb - XLSX.xlsx", filename)
                if suburb_name:
                    self.dataset[suburb_name.group(1)] = self.load_dataset(file_path)

    def list_suburbs(self, verbose=None):
        """List all the suburbs in the dataset."""
        verbose = self.get_verbose(verbose)
        self.load_all_datasets()
        self.suburbs = sorted(self.dataset.keys())

        if verbose:
            self.console.print("Suburbs List:", style="bold black")
            columns = Columns(
                [
                    f"{idx}: {suburb}"
                    for idx, suburb in enumerate(self.suburbs, start=1)
                ],
                width=25,
            )
            self.console.print(columns, style="bold cyan")

    def get_data(self, suburb_name: str) -> pd.DataFrame:
        """Get the data for a specific suburb as a DataFrame."""
        if suburb_name not in self.dataset:
            self.console.print(f"No data found for suburb: {suburb_name}")
            return

        data = self.dataset[suburb_name]
        df = pd.DataFrame(
            data,
            columns=["Category", "Subcategory", "Value", "Extra1", "Extra2"],
        )
        if df.empty:
            return

        df["Category"] = df["Category"].ffill()
        df.drop(columns=["Extra1", "Extra2"], inplace=True)
        return df

    def list_categories(self, verbose=None) -> pd.Series:
        """Get categories list"""
        verbose = self.get_verbose(verbose)
        suburb_df = self.get_data(self.suburbs[0])
        self.categories = suburb_df["Category"].unique()

        if verbose:
            self.console.print("Categories List:", style="bold black")
            for idx, category in enumerate(self.categories, start=1):
                self.console.print(f"{idx}. {category}", style="bold cyan")

    def get_category(self, category: str, suburb=None) -> pd.DataFrame:
        """Filter data for a specific category."""
        if suburb is None:
            suburb = self.suburbs[0]
        suburb_df = self.get_data(suburb)
        category_df = suburb_df[suburb_df["Category"] == category]
        return category_df

    def list_subcategories(
        self, category=None, console_print=True, verbose=None
    ) -> pd.Series:
        """Get subcategories list"""
        verbose = self.get_verbose(verbose)
        if not hasattr(self, "categories"):
            self.list_categories()

        def get(category):
            category_df = self.get_category(category)
            return sorted(category_df["Subcategory"].unique())

        if category is None:
            # List for all categories
            subcategories = []
            for category in self.categories:
                subcategories.extend(get(category))
        else:
            subcategories = get(category)

        if verbose:
            self.console.print("Subcategories List:", style="bold black")
            for idx, subcategory in enumerate(subcategories, start=1):
                if console_print:
                    self.console.print(f"{idx}. {subcategory}", style="bold cyan")
                else:
                    print(subcategory)

        return subcategories

    def get_category_across_all_suburbs(self, category: str) -> pd.DataFrame:
        """Get data for a specific category across all suburbs."""

        data = []
        for suburb in self.suburbs:
            category_df = self.get_category(category, suburb)
            category_df = category_df.drop(columns=["Category"])
            category_df["Suburb"] = suburb
            data.append(category_df)

        data = pd.concat(data, ignore_index=False)
        data.set_index(["Suburb", "Subcategory"], inplace=True)
        data = data["Value"].unstack(level="Subcategory")
        data = data.sort_index()
        data.columns.name = category
        data = data.stack().unstack(level="Suburb")
        return data

    def get_categories_across_all_suburbs(self, categories: List[str]) -> pd.DataFrame:
        """Get data for a list of categories across all suburbs."""
        data = []
        for category in categories:
            category_df = self.get_category_across_all_suburbs(category)
            data.append(category_df)
        combined_data = pd.concat(data, axis=1, keys=categories)
        return combined_data

    def get_subcategory_across_all_suburbs(self, subcategory: str) -> pd.DataFrame:
        """Get data for a specific subcategory across all suburbs."""

        data = []
        for suburb in self.suburbs:
            suburb_df = self.get_data(suburb)
            subcategory_df = suburb_df[suburb_df["Subcategory"] == subcategory]
            df = subcategory_df.drop(columns=["Subcategory"])
            df["Suburb"] = suburb
            data.append(df)

        data = pd.concat(data, ignore_index=True)
        data = data.drop(columns=["Category"])
        data.set_index("Suburb", inplace=True)
        data.columns = [subcategory]
        data = data.sort_index()
        return data

    def get_subcategories_across_all_suburbs(
        self, subcategories: List[str]
    ) -> pd.DataFrame:
        """Get data for a list of subcategories across all suburbs."""
        data = []
        for subcategory in subcategories:
            subcategory_df = self.get_subcategory_across_all_suburbs(subcategory)
            data.append(subcategory_df)

        data = pd.concat(data, axis=1)
        return data


class DataProcessor:
    def __init__(self, df: pd.DataFrame, random_state=42) -> None:
        self.df = df
        self.random_state = random_state

    def normalize(self):
        """Normalize the data"""
        df = self.df
        df_norm = minmax_scale(df, axis=0)
        df_norm = pd.DataFrame(df_norm, columns=df.columns, index=df.index)
        self.df = df_norm

    def get_correlation_matrix(self):
        """Get the correlation matrix"""
        return self.df.corr()

    def get_topk_abs_correlations(self, k=None):
        """Get the top k absolute correlations"""
        matrix = self.get_correlation_matrix()
        pairs = matrix.unstack().reset_index()
        pairs.columns = ["Variable1", "Variable2", "Correlation"]
        pairs["AbsCorrelation"] = pairs["Correlation"].abs()
        pairs = pairs[pairs["Variable1"] != pairs["Variable2"]]
        pairs = pairs.sort_values(by="AbsCorrelation", ascending=True)
        pairs = pairs[pairs["Variable1"] < pairs["Variable2"]]
        pairs = pairs.reset_index(drop=True)
        if k:
            pairs = pairs.head(k)
        return pairs

    def run_anova_analysis(self, subcategories):
        """Run ANOVA analysis"""
        df = self.df
        anova_results = {}
        for subcategory in subcategories:
            anova_results[subcategory] = f_oneway(
                *[df[subcategory].values for suburb in df.index]
            )
        return anova_results

    def plot_dendrogram(self, df=None):
        """Get the dendrogram plot"""
        if df is None:
            df = self.df
        linkage_matrix = sch.linkage(df, method="ward")
        plt.figure(figsize=(10, 7))
        dendrogram = sch.dendrogram(linkage_matrix, labels=df.index.tolist())
        plt.title("Dendrogram")
        plt.xlabel("Suburbs")
        plt.ylabel("Euclidean distances")
        plt.xticks(rotation=90)
        plt.show()

    def get_similarity_matrix(self, metric):
        """Get the similarity matrix"""

        if metric == "cosine":
            metric = cosine_similarity
        elif metric == "euclidean":
            metric = lambda x: 1 / (1 + euclidean_distances(x))

        df = self.df
        return pd.DataFrame(metric(df), index=df.index, columns=df.index)

    def get_similar_suburbs(self, similarity_matrix, n_neighbours=5):
        similarity_matrix = similarity_matrix.copy()
        np.fill_diagonal(similarity_matrix.values, 0)
        similar_suburbs = similarity_matrix.apply(
            lambda x: x.nlargest(n_neighbours).index.tolist(), axis=1
        )
        return similar_suburbs

    def plot_heatmap(self, matrix, title="Heatmap"):
        """Plot heatmap"""
        plt.figure(figsize=(10, 7))
        sns.heatmap(matrix, cmap="viridis", annot=False)
        plt.title(title)
        plt.show()

    def run_metric_mds_and_plot(
        self,
        similarity_matrix,
        n_components=2,
        isMetricMDS: bool = True,
    ):
        mds = MDS(
            n_components=n_components,
            metric=isMetricMDS,
            dissimilarity="precomputed",
            random_state=self.random_state,
        )
        dissimilarity_matrix = 1 - similarity_matrix
        mds_results = mds.fit_transform(dissimilarity_matrix)
        mds_df = pd.DataFrame(
            mds_results,
            index=similarity_matrix.index,
            columns=[f"MDS{i+1}" for i in range(n_components)],
        )

        plt.figure(figsize=(10, 7))
        sns.scatterplot(x="MDS1", y="MDS2", data=mds_df)
        for i in mds_df.index:
            plt.text(mds_df.loc[i, "MDS1"], mds_df.loc[i, "MDS2"], i, fontsize=9)
        plt.title("MDS Plot")
        plt.xlabel("MDS1")
        plt.ylabel("MDS2")
        plt.show()

        return mds_df


class LocationProcessor:
    def __init__(self, location_df: pd.DataFrame) -> None:
        self.location_df = location_df
        self.coordinates = None

    def extract_coordinates(self, location):
        match = re.search(r"(\d+)km ([A-Z]+) of Melbourne", location)
        if match:
            distance = int(match.group(1))
            direction = match.group(2)
            angle = {
                "N": 0,
                "NNE": 22.5,
                "NE": 45,
                "ENE": 67.5,
                "E": 90,
                "ESE": 112.5,
                "SE": 135,
                "SSE": 157.5,
                "S": 180,
                "SSW": 202.5,
                "SW": 225,
                "WSW": 247.5,
                "W": 270,
                "WNW": 292.5,
                "NW": 315,
                "NNW": 337.5,
            }[direction]
            return distance, angle
        return None, None

    def polar_to_cartesian(self, distance, angle):
        angle_rad = radians(angle)
        x = distance * cos(angle_rad)
        y = distance * sin(angle_rad)
        return x, y

    def get_coordinates(self):
        coordinates = {}
        for suburb, row in self.location_df.iterrows():
            distance, angle = self.extract_coordinates(row["Location"])
            if distance is not None and angle is not None:
                coordinates[suburb] = self.polar_to_cartesian(distance, angle)
        self.coordinates = coordinates
        return self.coordinates

    def calculate_proximity_matrix(self):
        coordinates = self.coordinates
        suburbs = list(coordinates.keys())
        proximity_matrix = pd.DataFrame(index=suburbs, columns=suburbs)
        for suburb1 in suburbs:
            for suburb2 in suburbs:
                x1, y1 = coordinates[suburb1]
                x2, y2 = coordinates[suburb2]
                distance = sqrt((x2 - x1) ** 2 + (y2 - y1) ** 2)
                proximity_matrix.loc[suburb1, suburb2] = distance
        return proximity_matrix

    def rotate_coordinates(self, coordinates, angle):
        angle_rad = radians(angle)
        rotated_coords = {}
        for suburb, (x, y) in coordinates.items():
            x_rot = x * cos(angle_rad) - y * sin(angle_rad)
            y_rot = x * sin(angle_rad) + y * cos(angle_rad)
            rotated_coords[suburb] = (x_rot, y_rot)
        return rotated_coords

    def reflect_coordinates(self, coordinates):
        reflected_coords = {}
        for suburb, (x, y) in coordinates.items():
            reflected_coords[suburb] = (x, -y)
        return reflected_coords

    def plot_coordinates(self, coordinates):
        plt.figure(figsize=(10, 10))
        for suburb, (x, y) in coordinates.items():
            plt.scatter(x, y, label=suburb)
            plt.text(x, y, suburb, fontsize=9)

        plt.title("Suburb Coordinates")
        plt.xlabel("X Coordinate")
        plt.ylabel("Y Coordinate")
        plt.grid(True)
        # plt.legend(loc='best', bbox_to_anchor=(1, 1))
        plt.show()
